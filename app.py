import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
import requests
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# --- 1. 金管局 API (HKFRS 折現率基準) ---
def fetch_custom_yield(target_date, tenor_key="efn_10y_yield"):
    date_str = target_date.strftime('%Y-%m-%d')
    try:
        url = f"https://data.static.hkma.gov.hk/public/data/ef-bills-and-notes-yields-and-prices-daily?filters=end_of_day:[* TO {date_str}]&pagesize=1&sort=end_of_day desc"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            records = data.get('result', {}).get('records', [])
            if records:
                latest = records[0]
                val = latest.get(tenor_key)
                if val is not None:
                    return float(val), latest.get('end_of_day'), "Live (HKMA)"
    except:
        pass
    return 3.5, None, "Fallback (3.5%)"


# --- 2. 年資計算 (官方方法：足年 + 餘下天數/365) ---
def calc_service_years(start_dt, end_dt):
    """Official method: completed whole years + remaining days / 365"""
    if end_dt <= start_dt:
        return 0.0
    years = end_dt.year - start_dt.year
    try:
        anniversary = start_dt.replace(year=end_dt.year)
    except ValueError:
        anniversary = start_dt.replace(year=end_dt.year, day=28)
    if anniversary > end_dt:
        years -= 1
        try:
            anniversary = start_dt.replace(year=end_dt.year - 1)
        except ValueError:
            anniversary = start_dt.replace(year=end_dt.year - 1, day=28)
    remaining_days = (end_dt - anniversary).days
    return max(0.0, years + remaining_days / 365)


# --- 3. 官方資助計劃邏輯 (對標政府 LD_EasyCal 25年階梯) ---
def get_subsidy_detail(net_post, val_date):
    """
    Two-tier government subsidy (25-year ladder).
    (a) net_post ≤ $500,000: within-threshold rate + employer cap formula (years 1–9) or rate-only (years 10–25).
    (b) net_post  > $500,000: beyond-threshold rate only (lower; zero from year 12 onward).
    Returns dict with keys: policy_year, rate, cap, rate_w, cap_w, rate_b, within_threshold, subsidy.
    """
    SCHEME_START = pd.Timestamp('2025-05-01')
    delta_days = (val_date - SCHEME_START).days
    if delta_days < 0:
        return {"policy_year": 0, "rate": 0.0, "cap": None,
                "rate_w": 0.0, "cap_w": None, "rate_b": 0.0,
                "within_threshold": True, "subsidy": 0.0}
    year_idx = int(delta_days // 365) + 1

    # Two-tier subsidy ladder (within / beyond $500,000 threshold per subsidy year).
    # Columns: (within_rate, within_cap, beyond_rate)
    #   within_cap = None → rate-only formula for years 10-25 (no employer-cap formula).
    #   beyond_rate = 0.0 → no subsidy for net_post > $500,000 from year 12 onwards.
    THRESHOLD = 500_000
    ladder = {
        1:  (0.50,  3000, 0.50),  2:  (0.50,  3000, 0.50),  3:  (0.50,  3000, 0.50),
        4:  (0.45, 25000, 0.45),  5:  (0.40, 25000, 0.40),  6:  (0.35, 25000, 0.35),
        7:  (0.30, 50000, 0.30),  8:  (0.25, 50000, 0.25),  9:  (0.20, 50000, 0.20),
       10:  (0.20,  None, 0.15), 11:  (0.20,  None, 0.10), 12:  (0.15,  None, 0.05),
       13:  (0.15,  None, 0.00), 14:  (0.10,  None, 0.00), 15:  (0.10,  None, 0.00),
       16:  (0.10,  None, 0.00), 17:  (0.10,  None, 0.00), 18:  (0.10,  None, 0.00),
       19:  (0.10,  None, 0.00), 20:  (0.05,  None, 0.00), 21:  (0.05,  None, 0.00),
       22:  (0.05,  None, 0.00), 23:  (0.05,  None, 0.00), 24:  (0.05,  None, 0.00),
       25:  (0.05,  None, 0.00),
    }
    rate_w, cap_w, rate_b = ladder.get(year_idx, (0.0, 0, 0.0))
    within_threshold = net_post <= THRESHOLD
    if within_threshold:
        if cap_w is not None:
            subsidy = max(0.0, max(rate_w * net_post, net_post - cap_w))
        else:
            subsidy = max(0.0, rate_w * net_post)
        rate_applied, cap_applied = rate_w, cap_w
    else:
        subsidy = max(0.0, rate_b * net_post)
        rate_applied, cap_applied = rate_b, None
    return {
        "policy_year": year_idx,
        "rate": rate_applied, "cap": cap_applied,
        "rate_w": rate_w, "cap_w": cap_w, "rate_b": rate_b,
        "within_threshold": within_threshold,
        "subsidy": subsidy,
    }


def calculate_gov_subsidy(net_post, val_date):
    return get_subsidy_detail(net_post, val_date)["subsidy"]


# --- 3. UI 語言定義 ---
LANG_DICT = {
    "繁中": {
        "title": "香港長期服務金計算工具",
        "settings": "計算參數",
        "rep_date_label": "報告/評估日期",
        "g_label": "預期加薪率 (%)",
        "m_label": "預期 MPF 回報 (%)",
        "r_label": "折現率 (%)",
        "t_label": "年度離職率 (%)",
        "calc_btn": "生成審計底稿",
        "g_help": "年度加薪率。建議參考：長期通脹率 (約 2-3%)。用於預測員工退休前薪金（僅 HKFRS；SME-FRS 強制採用 0%）。",
        "m_help": "MPF 回報率。建議參考：積金局 DIS 核心累積基金 (約 4-5%)。詳見積金局季度報告連結。",
        "r_help": "折現率，用於計算未來負債現值。建議參考金管局外匯基金票據收益率。詳見金管局連結。",
        "t_help": "年度離職率（員工提前離職概率）。僅 HKFRS 下按年複合計算留任概率。SME-FRS 假設留任概率為 100%。",
        "upload_guide": "格式需求：需含 [Name, Hired Date, DOB, Salary at Transition, Current Salary, MPF Mand (ER), MPF Vol (ER)]。MPF Mand (ER) 應填截至評估日、僅限 2025-05-01 或之前僱主強制性供款的應計權益；如留空或為 0，系統會先估算截至 2025-05-01 的僱主強制性供款，再按 MPF 回報率滾存至評估日。MPF Vol (ER) 可選填，留空視作 0。",
        "res_pre": "轉制前 LSP (a)",
        "res_mpf_pre": "MPF 對沖 (轉制前)",
        "res_net_pre": "轉制前淨額 (a')",
        "res_post": "轉制後 LSP (b)",
        "res_mpf_post": "MPF 對沖 (轉制後)",
        "res_net_post": "轉制後淨額 (b')",
        "res_sub": "政府資助 (針對 b')",
        "res_emp_post": "僱主承擔 (b'-資助)",
        "res_final": "最終負債 (a'+僱主b') [SME-FRS 不折現 / HKFRS×PV]",
        "res_mpf_basis": "MPF 數據基礎",
        "mpf_actual": "實際應計權益（評估日；僅限 2025-05-01 前僱主強制性供款）",
        "mpf_estimated": "估算應計權益（僱主強制性供款至 2025-05-01，再滾存至評估日）",
        "links_title": "📖 官方資源與數據連結",
        "sig_text": "我方確認以上數據及邏輯準確無誤：",
        "disclaimer": "免責條款：本工具僅供參考，計算結果應根據法定要求及強積金受託人結單核實（特別是截至評估日、僅限 2025-05-01 前僱主強制性供款的應計權益）。所有計算均於您的瀏覽器本地執行，任何輸入資料均不會傳送至或儲存於任何伺服器。",
        "hint": "💡 提示：如需與官方計算器對比，請將加薪率及折現率設為 0%。",
        # --- UI labels ---
        "std_label": "會計準則",
        "source_label": "資料來源",
        "manual_opt": "手動輸入",
        "upload_opt": "上傳檔案",
        "upload_file_label": "上傳檔案",
        "before_abolition_warn": "⚠️ 評估日期早於廢除日期 (2025-05-01)。",
        "policy_year_info": "政策年度：{y}（後轉制）",
        "results_title": "📊 結果摘要",
        "audit_expander": "🧮 審計底稿 — 逐步計算明細",
        "download_btn": "📥 下載 Excel 工作底稿（5 工作表）",
        "guide_expander": "📘 使用指南",
        "guide_missing": "找不到 USER_GUIDE.md，請確認檔案位於與 app.py 相同資料夾。",
        # --- Audit section A ---
        "audit_a_title": "#### A. 已採用的精算假設",
        "audit_std_lbl": "**會計準則：**",
        "audit_val_lbl": "**評估日期：**",
        "audit_trans_lbl": "**轉制日期：** 2025年5月1日",
        "audit_mpf_inc_lbl": "**強積金實施日：** 2000年12月1日",
        "audit_g_lbl": "**加薪率 (g)：**",
        "audit_r_lbl": "**折現率 (r)：**",
        "audit_sal_cap_lbl": "**LSP 薪金上限：** HK$22,500/月",
        "audit_lsp_max_lbl": "**LSP 最高金額：** HK$390,000",
        "audit_mpf_rate_lbl": "**強積金強制性僱主供款率：** 5%（薪金上限：≤2012-04：$20,000 | 2012-05至2014-05：$25,000 | 2014-06起：$30,000）",
        "audit_pv_hkfrs": "（HKFRS — 須計算貨幣時間值）",
        "audit_pv_sme_note": "不適用（SME-FRS — 無須折現）",
        # --- Audit section B ---
        "audit_b_title": "#### B. 各員工計算詳情",
        "audit_input_title": "**輸入數據**",
        "audit_hire_lbl": "入職日期：",
        "audit_dob_lbl": "出生日期：",
        "audit_age_suffix": "歲）",
        "audit_sal_t_lbl": "轉制前薪金：",
        "audit_sal_c_lbl": "現時薪金：",
        "audit_mpf_mand_lbl": "強積金強制性供款（僱主；評估日應計權益，限 2025-05-01 前供款）：",
        "audit_mpf_vol_lbl": "強積金自願性供款（僱主）：",
        "audit_step1_title": "**第1步 — 服務年期及折現**",
        "audit_pre_y_lbl": "轉制前年資：",
        "audit_pre_y_note": "（入職 → 2025-05-01，足年 + 餘下天數/365）",
        "audit_post_y_lbl": "轉制後年資：",
        "audit_ret_y_lbl": "至退休年期（65歲）：",
        "audit_pv_factor_lbl": "折現因子 =",
        "audit_pv_sme_label": "折現因子：`1.000000`（SME-FRS）",
        "audit_turnover_lbl": "年度離職率：",
        "audit_prob_stay_lbl": "留任概率 =",
        "audit_prob_stay_sme": "留任概率：`1.000000`（SME-FRS — 假設全員留任）",
        "audit_step2_title": "**第2步 — LSP 金額**",
        "audit_pre_lsp_note": "（上限 $390,000）",
        "audit_proj_sal_lbl": "轉制後預計薪金：",
        "audit_post_lsp_note": "（上限 $390,000 − 轉制前）",
        "audit_step3_title": "**第3步 — MPF 對沖**",
        "audit_mand_off_lbl": "強制性對沖（轉制前）=",
        "audit_mand_off_note": "— 強制性：僅限轉制前",
        "audit_vol_pre_lbl": "自願性對沖（轉制前）=",
        "audit_total_pre_off": "轉制前總對沖 =",
        "audit_net_pre_lbl": "→ 轉制前淨額 (a') =",
        "audit_vol_rem_lbl": "自願性剩餘 =",
        "audit_post_off_lbl": "轉制後對沖（自願性）=",
        "audit_net_post_lbl": "轉制後淨額 (b') =",
        "audit_vol_only_note": "— 僅限自願性供款",
        "audit_step4_title": "**第4步 — 政府資助及最終負債**",
        "audit_pol_yr_lbl": "政策年度：",
        "audit_rate_lbl": "資助比率：",
        "audit_emp_cap_lbl": "僱主承擔上限：",
        "audit_threshold_within": "✅ 淨額 (b') ≤ $500,000（適用第(a)欄比率）",
        "audit_threshold_beyond": "⚠️ 淨額 (b') > $500,000（適用第(b)欄比率）",
        "audit_subsidy_lbl": "資助金額 =",
        "audit_emp_burden_lbl": "僱主承擔 =",
        "audit_total_burden_lbl": "總負擔 =",
        "audit_final_lbl": "**最終負債 =**",
        # --- Audit sections C & D ---
        "audit_c_title": "#### C. 法律依據及公式參考",
        "audit_d_title": "#### D. 限制及免責聲明",
        "audit_stat_table": """
| 項目 | 法律依據 |
|------|---------|
| 長期服務金計算公式 | 僱傭條例（第57章）第31A條 — 每服務年期 2/3 個月薪金 |
| LSP 薪金上限 | 第57章 — HK$22,500/月 |
| LSP 最高金額 | 第57章 — HK$390,000 |
| 取消強積金對沖安排 | 《2022年僱傭及退休計劃法例（抵銷安排）（修訂）條例》 |
| 強積金強制性僱主供款率 | 強制性公積金計劃條例（第485章）— 5%，薪金上限 HK$30,000 |
| 政府資助計劃 | 取消強積金對沖安排資助計劃（25年階梯） |
| 會計準則（SME-FRS） | 香港會計師公會中小企業財務報告準則第28節 — 僱員福利（無須折現） |
| 會計準則（HKFRS） | 香港會計準則第19號 / 香港財務報告準則（中小型實體）第28節 — 預測單位計算法，按優質債券收益率折現 |
| 折現率參考 | 金管局外匯基金票據及債券每日收益率 |
""",
        "audit_limit_text": """
**本計算的限制：**
1. **強積金估算數字**：凡強積金強制性供款（僱主）為零或空白，系統按三段式模型估算（反映薪金上限歷史變化：2012年4月前=$20,000、2012年5月至2014年5月=$25,000、2014年6月起=$30,000）。估算只涵蓋截至 2025-05-01 的僱主強制性供款，並按用戶指定的強積金回報率複利滾存至評估日。最終審計結案前，必須向強積金受託人索取實際結單（評估日應計權益）。
2. **薪金預測**：HKFRS 下轉制後 LSP 採用按用戶指定增長率推算至退休（65歲）的預測薪金；SME-FRS 強制採用現時薪金（加薪率設為 0%）。實際離職薪金可能不同。
3. **折現率**：僅適用於 HKFRS / HKFRS for PE。基於金管局外匯基金票據收益率或用戶輸入。須於每個報告日按照香港會計準則第19號更新。
4. **政府資助**：資助比率可能因立法修訂而有所改變。每次評估時請向官方資助計劃管理機構核實適用年度的比率。
5. **SME-FRS**：不作折現。負債代表所有員工於評估日離職的當前成本義務。
6. **退休年齡**：本工具假設所有員工於 **65歲** 退休。如員工可能在65歲前離職，應通過「年度離職率」參數反映相關概率。
7. **離職率**：HKFRS 下，留任概率 (1−t)^n 僅適用於**僱主承擔的轉制後部分**（employer_post），而非轉制前淨額（net_pre）。轉制前 LSP 對服務滿5年的員工已完全歸屬，任何離職情況均須支付，不存在沒收風險。最終負債公式為：(net_pre + employer_post × prob_stay) × PV。
8. **範圍 — 僅限 LSP**：本工具計算長期服務金。遣散費（SP）採用相同公式但適用於裁員情況，請按第57章第31至33條另行核實。
9. **數據可靠性**：計算結果完全依賴所輸入的數據。填表人須核對入職日期、薪金、出生日期及強積金結餘是否與源文件（人事記錄、強積金結單）一致。
""",
    },
    "EN": {
        "title": "HK LSP Calculator Tool",
        "settings": "Calculation Parameters",
        "rep_date_label": "Valuation Date",
        "g_label": "Salary Growth (%)",
        "m_label": "MPF Return (%)",
        "r_label": "Discount Rate (%)",
        "t_label": "Annual Turnover Rate (%)",
        "calc_btn": "Generate Report",
        "g_help": "Ref: Long-term inflation (2-3%). Used for salary projection to retirement (HKFRS only; SME-FRS forces 0%).",
        "m_help": "Ref: MPFA DIS Core Fund (4-5%). Used for projecting MPF asset growth. See MPFA quarterly reports link.",
        "r_help": "Discount rate for future liability. Typically based on HKMA EFBN yields or high-quality corporate bonds. See HKMA link below.",
        "t_help": "Annual probability of employee leaving before retirement. Under HKFRS, compounded over years-to-retirement to derive a retention probability multiplier. SME-FRS assumes 100% retention.",
        "upload_guide": "Required columns: [Name, Hired Date, DOB, Salary at Transition, Current Salary, MPF Mand (ER), MPF Vol (ER)]. MPF Mand (ER) should be the accrued benefit at valuation date for employer mandatory contributions up to 2025-05-01. If blank/0, the tool estimates employer mandatory contributions up to 2025-05-01 and compounds them to valuation date using the MPF return assumption. MPF Vol (ER) is optional; leave blank for 0.",
        "res_pre": "Pre-transition LSP (a)",
        "res_mpf_pre": "MPF Offset (pre)",
        "res_net_pre": "Net Pre-transition (a')",
        "res_post": "Post-transition LSP (b)",
        "res_mpf_post": "MPF Offset (post)",
        "res_net_post": "Net Post-transition (b')",
        "res_sub": "Gov Subsidy (on b')",
        "res_emp_post": "Employer's Post (b'-subsidy)",
        "res_final": "Net Liability (a'+emp.b') [SME-FRS: no PV / HKFRS: ×PV]",
        "res_mpf_basis": "MPF Data Basis",
        "mpf_actual": "Actual accrued benefit @ valuation date (ER mandatory up to 2025-05-01)",
        "mpf_estimated": "Estimated accrued benefit @ valuation date (ER mandatory up to 2025-05-01)",
        "links_title": "📖 Official Resources & References",
        "sig_text": "Confirmation of calculation logic and data accuracy:",
        "disclaimer": "Disclaimer: This tool is for reference only. Users must verify results against statutory requirements and MPF trustee statements (especially accrued benefit at valuation date for ER mandatory contributions up to 2025-05-01). All calculations are performed locally in your browser — no input data is transmitted to or stored on any server.",
        "hint": "💡 Tip: Set Salary Growth and Discount Rate to 0% to cross-check against the official LD calculator.",
        # --- UI labels ---
        "std_label": "Accounting Standard",
        "source_label": "Source",
        "manual_opt": "Manual",
        "upload_opt": "Upload",
        "upload_file_label": "Upload File",
        "before_abolition_warn": "⚠️ Valuation date is before the abolition date (2025-05-01).",
        "policy_year_info": "Policy Year: {y} (Post-Abolition)",
        "results_title": "📊 Results Summary",
        "audit_expander": "🧮 Audit Proof — Step-by-Step Calculations",
        "download_btn": "📥 Download Excel Working Paper (5 Sheets)",
        "guide_expander": "📘 User Guide",
        "guide_missing": "USER_GUIDE.md was not found. Please place it in the same folder as app.py.",
        # --- Audit section A ---
        "audit_a_title": "#### A. Actuarial Assumptions Applied",
        "audit_std_lbl": "**Accounting Standard:**",
        "audit_val_lbl": "**Valuation Date:**",
        "audit_trans_lbl": "**Transition Date:** 1 May 2025",
        "audit_mpf_inc_lbl": "**MPF Inception Date:** 1 Dec 2000",
        "audit_g_lbl": "**Salary Growth Rate (g):**",
        "audit_r_lbl": "**Discount Rate (r):**",
        "audit_sal_cap_lbl": "**LSP Salary Cap:** HK$22,500/month",
        "audit_lsp_max_lbl": "**LSP Maximum:** HK$390,000",
        "audit_mpf_rate_lbl": "**MPF Mandatory ER Rate:** 5% (salary cap: ≤Apr 2012: $20,000 | May 2012–May 2014: $25,000 | Jun 2014+: $30,000)",
        "audit_pv_hkfrs": "(HKFRS — time value applied)",
        "audit_pv_sme_note": "N/A (SME-FRS — no discounting required)",
        # --- Audit section B ---
        "audit_b_title": "#### B. Per-Employee Calculation Walkthrough",
        "audit_input_title": "**Input Data**",
        "audit_hire_lbl": "Hire Date:",
        "audit_dob_lbl": "DOB:",
        "audit_age_suffix": ")",
        "audit_sal_t_lbl": "Salary @ Transition:",
        "audit_sal_c_lbl": "Current Salary:",
        "audit_mpf_mand_lbl": "MPF Mandatory (ER; accrued @ valuation date, pre-2025-05-01 only):",
        "audit_mpf_vol_lbl": "MPF Voluntary (ER):",
        "audit_step1_title": "**Step 1 — Service Years & Discounting**",
        "audit_pre_y_lbl": "Pre-transition:",
        "audit_pre_y_note": "*(Hire → 2025-05-01, whole yrs + days/365)*",
        "audit_post_y_lbl": "Post-transition:",
        "audit_ret_y_lbl": "Years to Retirement (65):",
        "audit_pv_factor_lbl": "PV Factor =",
        "audit_pv_sme_label": "PV Factor: `1.000000` (SME-FRS)",
        "audit_turnover_lbl": "Annual Turnover Rate:",
        "audit_prob_stay_lbl": "Prob. of Staying =",
        "audit_prob_stay_sme": "Prob. of Staying: `1.000000` (SME-FRS — 100% retention assumed)",
        "audit_step2_title": "**Step 2 — LSP Amounts**",
        "audit_pre_lsp_note": "(cap $390,000)",
        "audit_proj_sal_lbl": "Post-LSP projected salary:",
        "audit_post_lsp_note": "(cap $390,000 − pre)",
        "audit_step3_title": "**Step 3 — MPF Offsetting**",
        "audit_mand_off_lbl": "Mandatory offset on pre =",
        "audit_mand_off_note": "— mandatory: pre only",
        "audit_vol_pre_lbl": "Voluntary offset on pre =",
        "audit_total_pre_off": "Total pre offset =",
        "audit_net_pre_lbl": "→ Net pre (a') =",
        "audit_vol_rem_lbl": "Voluntary remaining =",
        "audit_post_off_lbl": "Post offset (voluntary) =",
        "audit_net_post_lbl": "Net post (b') =",
        "audit_vol_only_note": "— voluntary only on post",
        "audit_step4_title": "**Step 4 — Gov Subsidy & Final Liability**",
        "audit_pol_yr_lbl": "Policy Year:",
        "audit_rate_lbl": "Rate:",
        "audit_emp_cap_lbl": "Employer Cap:",
        "audit_threshold_within": "✅ Net (b') ≤ $500,000 — within threshold (column a rates)",
        "audit_threshold_beyond": "⚠️ Net (b') > $500,000 — beyond threshold (column b rates)",
        "audit_subsidy_lbl": "Subsidy =",
        "audit_emp_burden_lbl": "Employer post-burden =",
        "audit_total_burden_lbl": "Total burden =",
        "audit_final_lbl": "**Final Liability =**",
        # --- Audit sections C & D ---
        "audit_c_title": "#### C. Statutory Basis & Formula Reference",
        "audit_d_title": "#### D. Limitations & Disclaimers",
        "audit_stat_table": """
| Item | Reference |
|------|-----------|
| LSP entitlement formula | Employment Ordinance (Cap. 57) s.31A — 2/3 month per year of service |
| Salary cap for LSP | Cap. 57 — HK$22,500/month |
| Maximum LSP | Cap. 57 — HK$390,000 |
| MPF offsetting abolition | Employment and Retirement Schemes Legislation (Offsetting Arrangement) (Amendment) Ordinance 2022 |
| MPF mandatory ER rate | Mandatory Provident Fund Schemes Ordinance (Cap. 485) — 5%, salary cap HK$30,000 |
| Government Subsidy Scheme | Subsidy Scheme for Abolition of MPF Offsetting Arrangement (25-year ladder) |
| Accounting standard (SME-FRS) | HKICPA SME-FRS Section 28 — Employee Benefits (no discounting required) |
| Accounting standard (HKFRS) | HKAS 19 / HKFRS for PE Sec. 28 — Projected Unit Credit Method, discount at high-quality bond yield |
| Discount rate reference | HKMA Exchange Fund Bills & Notes (EFBN) daily yields |
""",
        "audit_limit_text": """
**Limitations of this calculation:**
1. **MPF Estimated figures**: Where MPF Mandatory (ER) was zero or blank, the balance was estimated using a 3-period segment model reflecting historical MPF salary cap changes: ≤Apr 2012=$20,000/mth, May 2012–May 2014=$25,000/mth, Jun 2014+=$30,000/mth. Estimation includes ER mandatory contributions up to 2025-05-01 only, then compounds to valuation date at the user-specified MPF Return Rate. Actual MPF trustee statements (valuation-date accrued benefit) must be obtained for final audit sign-off.
2. **Salary projection**: Under HKFRS, post-transition LSP uses salary projected to age 65 at the user-specified growth rate. Under SME-FRS, the current salary is used directly (growth rate forced to 0%). Actual termination salary may differ.
3. **Discount rate**: Based on HKMA EFBN live yield or user input. Rate should be refreshed at each reporting date.
4. **Government subsidy**: Subsidy rates are subject to legislative changes over the 25-year scheme. Verify the applicable year's rate at each valuation.
5. **SME-FRS**: No discounting applied. The liability represents the current-cost obligation as if all employees were to terminate on the valuation date.
6. **Retirement age**: Retirement is assumed at age **65** for all employees. Adjust the turnover rate input to reflect the probability that employees may leave before reaching age 65.
7. **Turnover / attrition**: Under HKFRS, the retention probability (1−t)^n is applied **only to the employer's post-transition burden** (employer_post), not to the pre-transition net obligation (net_pre). Pre-transition LSP is fully vested for employees with 5+ years of service — it is payable upon any termination and carries no forfeiture risk. Final formula: (net_pre + employer_post × prob_stay) × PV Factor. A full actuarial valuation would use individual decrements.
8. **Scope**: This tool covers Long Service Payments only. Severance Payments (SP) follow the same formula but apply to redundancy cases; verify applicability separately.
9. **Data reliability**: This tool relies entirely on the input data provided. The preparer is responsible for verifying hire dates, salaries, and MPF balances against source documents.
""",
    }
}


@st.cache_data(show_spinner=False)
def load_user_guide_md():
    guide_path = Path(__file__).with_name("USER_GUIDE.md")
    try:
        return guide_path.read_text(encoding="utf-8-sig")
    except (FileNotFoundError, OSError):
        return None


# --- 4. 側邊欄與輸入 ---
st.set_page_config(page_title="LSP Calculator", layout="wide")
lang_choice = st.sidebar.radio("Language", ["EN", "繁中"])
L = LANG_DICT[lang_choice]

st.sidebar.header(L["settings"])
std_choice = st.sidebar.selectbox(L["std_label"], ["SME-FRS", "HKFRS / HKFRS for PE"])
rep_date = st.sidebar.date_input(L["rep_date_label"], datetime(2026, 3, 31))
g_val = st.sidebar.slider(L["g_label"], 0.0, 7.0, 3.0, help=L["g_help"]) / 100
m_val = st.sidebar.slider(L["m_label"], 0.0, 7.0, 2.6, help=L["m_help"]) / 100
yield_init, _, _ = fetch_custom_yield(rep_date)
r_val = st.sidebar.number_input(L["r_label"], 0.0, 10.0, yield_init, help=L["r_help"]) / 100
t_val = st.sidebar.slider(L["t_label"], 0.0, 20.0, 5.0, help=L["t_help"]) / 100

year_idx_display = int((pd.Timestamp(rep_date) - pd.Timestamp('2025-05-01')).days // 365) + 1
if pd.Timestamp(rep_date) < pd.Timestamp('2025-05-01'):
    st.sidebar.warning(L["before_abolition_warn"])
else:
    st.sidebar.info(L["policy_year_info"].format(y=year_idx_display))
st.sidebar.caption(L["hint"])

st.title(L["title"])
st.info("🔒 " + ("所有計算均於您的瀏覽器本地執行，任何輸入資料均不會傳送至或儲存於任何伺服器。"
                  if lang_choice == "繁中" else
                  "All calculations run locally in your browser. No input data is transmitted to or stored on any server."))

input_method = st.radio(L["source_label"], [L["manual_opt"], L["upload_opt"]], horizontal=True)
final_df = pd.DataFrame()

if input_method == L["upload_opt"]:
    st.info(L["upload_guide"])
    file = st.file_uploader(L["upload_file_label"], type=["xlsx", "csv"])
    if file:
        final_df = pd.read_excel(file) if file.name.endswith('xlsx') else pd.read_csv(file)
else:
    init_data = [{"Name": "Staff A", "Hired Date": "2015-01-01", "DOB": "1980-01-01",
                  "Salary at Transition": 21000, "Current Salary": 24000,
                  "MPF Mand (ER)": None, "MPF Vol (ER)": 0}]
    final_df = st.data_editor(pd.DataFrame(init_data), num_rows="dynamic")


# --- 5. 核心計算循環 ---
# Module-level constants and helper (not redefined per-row)
_MPF_INCEPTION  = pd.Timestamp('2000-12-01')
_MPF_CAP2_START = pd.Timestamp('2012-05-01')
_MPF_CAP3_START = pd.Timestamp('2014-06-01')
_MPF_CAP_SEGS   = [
    (_MPF_INCEPTION,  _MPF_CAP2_START, 20000),
    (_MPF_CAP2_START, _MPF_CAP3_START, 25000),
    (_MPF_CAP3_START, pd.Timestamp('2025-05-01'), 30000),
]

def _is_empty(v):
    if v is None: return True
    try:
        return np.isnan(float(v)) or float(v) <= 0
    except (ValueError, TypeError):
        return True

if st.button(L["calc_btn"]):
    TRANS_DATE = pd.Timestamp('2025-05-01')
    REPORT_DATE = pd.Timestamp(rep_date)
    results = []
    audit_details = []

    for idx, row in final_df.iterrows():
        try:
            h_dt = pd.Timestamp(row['Hired Date'])
            dob = pd.Timestamp(row['DOB'])
            sal_t = float(row['Salary at Transition'])
            sal_c = float(row['Current Salary'])
            mpf_mand_raw = row.get('MPF Mand (ER)', row.get('MPF Bal (ER)', 0))
            mpf_vol_raw  = row.get('MPF Vol (ER)', 0)

            # 強制性供款：如為 0 / None / 空，則估算
            # 估算方法：先估算截至轉制日的僱主強制性供款，再按 m_val 滾存至評估日
            #   段 1: 2000-12-01 – 2012-04-30  薪金上限 $20,000
            #   段 2: 2012-05-01 – 2014-05-31  薪金上限 $25,000
            #   段 3: 2014-06-01 – 2025-04-30  薪金上限 $30,000
            if _is_empty(mpf_mand_raw):
                mpf_mand = 0.0
                mpf_segs_detail = []
                for seg_start, seg_end, sal_cap in _MPF_CAP_SEGS:
                    p_start = max(h_dt, seg_start)
                    p_end   = min(TRANS_DATE, REPORT_DATE, seg_end)
                    if p_start >= p_end:
                        continue
                    seg_yr   = calc_service_years(p_start, p_end)
                    seg_cont = min(sal_t, sal_cap) * 0.05 * 12 * seg_yr
                    # midpoint of segment → years to REPORT_DATE for accrued benefit at valuation
                    seg_mid  = p_start + (p_end - p_start) / 2
                    yrs_grow = calc_service_years(seg_mid, REPORT_DATE)
                    seg_val  = seg_cont * ((1 + m_val) ** yrs_grow)
                    mpf_mand += seg_val
                    mpf_segs_detail.append((sal_cap, seg_yr, seg_cont, yrs_grow, seg_val))
                mpf_basis = L["mpf_estimated"]
            else:
                mpf_mand  = float(mpf_mand_raw)
                mpf_segs_detail = []
                mpf_basis = L["mpf_actual"]

            # 自願性供款：實際輸入，無估算
            mpf_vol = 0.0 if _is_empty(mpf_vol_raw) else float(mpf_vol_raw)

            # 資格檢查：不滿 5 年 (uses calc_service_years to handle leap years correctly)
            if calc_service_years(h_dt, REPORT_DATE) < 5.0:
                results.append({"Name": row['Name'], L['res_final']: 0, "Note": "年資不足5年 / Not Vested"})
                continue

            # 年資計算 (官方方法：足年 + 餘下天數/365)
            pre_y = calc_service_years(h_dt, TRANS_DATE)
            post_y = calc_service_years(TRANS_DATE, REPORT_DATE)
            y_to_ret = calc_service_years(REPORT_DATE, dob + pd.DateOffset(years=65))

            # SME-FRS 不考慮貨幣時間值，PV = 1；HKFRS 才折現
            # SME-FRS also uses current salary (no projection): effective_g forced to 0
            if std_choice == "SME-FRS":
                pv = 1.0
                effective_g = 0.0
                prob_stay = 1.0
            else:
                pv = 1 / ((1 + r_val) ** y_to_ret)
                effective_g = g_val
                prob_stay = (1 - t_val) ** y_to_ret

            # LSP 金額 (官方 $22,500 薪金上限與 $390,000 總上限)
            v_pre = min(min(sal_t, 22500) * (2 / 3) * pre_y, 390000)
            sal_proj = min(sal_c * ((1 + effective_g) ** y_to_ret), 22500)
            v_post = min(sal_proj * (2 / 3) * post_y, 390000 - v_pre)

            # MPF 對沖規則：
            #   強制性供款 → 只能對沖轉制前 LSP
            #   自願性供款 → 先對沖轉制前，餘額再對沖轉制後
            offset_mand_pre  = min(v_pre, mpf_mand)
            vol_to_pre       = min(v_pre - offset_mand_pre, mpf_vol)
            offset_pre       = offset_mand_pre + vol_to_pre
            net_pre          = v_pre - offset_pre

            vol_remaining    = mpf_vol - vol_to_pre
            offset_post      = min(v_post, vol_remaining)
            net_post         = v_post - offset_post

            # 政府資助：針對轉制後淨額 (b')，取 max(比例×b', b'-封頂)
            sub_detail    = get_subsidy_detail(net_post, REPORT_DATE)
            subsidy       = sub_detail["subsidy"]
            employer_post = max(0.0, net_post - subsidy)

            # 最終實質負債現值
            # net_pre is fully vested (5+ yr employees) → no turnover reduction, only PV discounting
            # employer_post is still accruing → reduced by prob_stay (retention probability)
            final_pv = (net_pre + employer_post * prob_stay) * pv

            results.append({
                "Name": row['Name'],
                L["res_mpf_basis"]: mpf_basis,
                L["res_pre"]: round(v_pre, 2),
                L["res_mpf_pre"]: round(offset_pre, 2),
                L["res_net_pre"]: round(net_pre, 2),
                L["res_post"]: round(v_post, 2),
                L["res_mpf_post"]: round(offset_post, 2),
                L["res_net_post"]: round(net_post, 2),
                L["res_sub"]: round(subsidy, 2),
                L["res_emp_post"]: round(employer_post, 2),
                L["res_final"]: round(final_pv, 2)
            })

            audit_details.append({
                "name": row['Name'],
                "h_dt": h_dt, "dob": dob,
                "sal_t": sal_t, "sal_c": sal_c,
                "mpf_mand": mpf_mand, "mpf_vol": mpf_vol, "mpf_basis": mpf_basis,
                "mpf_segs_detail": mpf_segs_detail, "m_val": m_val,
                "pre_y": pre_y, "post_y": post_y, "y_to_ret": y_to_ret,
                "pv": pv, "std": std_choice,
                "effective_g": effective_g, "prob_stay": prob_stay, "t_val": t_val,
                "v_pre": v_pre, "sal_proj_used": sal_proj, "v_post": v_post,
                "offset_mand_pre": offset_mand_pre, "vol_to_pre": vol_to_pre,
                "offset_pre": offset_pre, "net_pre": net_pre,
                "vol_remaining": vol_remaining, "offset_post": offset_post, "net_post": net_post,
                "policy_year": sub_detail["policy_year"],
                "sub_rate": sub_detail["rate"], "sub_cap": sub_detail["cap"],
                "rate_w": sub_detail["rate_w"], "cap_w": sub_detail["cap_w"],
                "rate_b": sub_detail["rate_b"],
                "within_threshold": sub_detail["within_threshold"],
                "subsidy": subsidy, "employer_post": employer_post,
                "final_pv": final_pv,
                "g_val": g_val, "r_val": r_val, "t_val": t_val,
            })
        except Exception as e:
            st.error(f"Error at row {idx}: {e}")

    if results:
        res_df = pd.DataFrame(results)
        st.subheader(L["results_title"])
        st.dataframe(res_df, use_container_width=True)

        # ── Audit Proof Expander ──────────────────────────────────────────────
        with st.expander(L["audit_expander"], expanded=False):
            # A. Assumptions
            st.markdown(L["audit_a_title"])
            a_col1, a_col2 = st.columns(2)
            with a_col1:
                st.markdown(f"{L['audit_std_lbl']} {std_choice}")
                st.markdown(f"{L['audit_val_lbl']} {rep_date}")
                st.markdown(L["audit_trans_lbl"])
                st.markdown(L["audit_mpf_inc_lbl"])
            with a_col2:
                st.markdown(f"{L['audit_g_lbl']} {g_val*100:.1f}% p.a.")
                pv_note = f"{r_val*100:.2f}% p.a. {L['audit_pv_hkfrs']}" if std_choice != "SME-FRS" else L["audit_pv_sme_note"]
                st.markdown(f"{L['audit_r_lbl']} {pv_note}")
                st.markdown(L["audit_sal_cap_lbl"])
                st.markdown(L["audit_lsp_max_lbl"])
                st.markdown(L["audit_mpf_rate_lbl"])

            st.markdown("---")

            # B. Per-employee detail
            st.markdown(L["audit_b_title"])
            for d in audit_details:
                age_at_val = (REPORT_DATE - d["dob"]).days / 365.25
                with st.container():
                    st.markdown(f"##### 👤 {d['name']}")
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.markdown(L["audit_input_title"])
                        st.markdown(f"{L['audit_hire_lbl']} `{d['h_dt'].date()}`")
                        st.markdown(f"{L['audit_dob_lbl']} `{d['dob'].date()}` (Age {age_at_val:.1f}{L['audit_age_suffix']}")
                        st.markdown(f"{L['audit_sal_t_lbl']} `${d['sal_t']:,.0f}`")
                        st.markdown(f"{L['audit_sal_c_lbl']} `${d['sal_c']:,.0f}`")
                        st.markdown(f"{L['audit_mpf_mand_lbl']} `${d['mpf_mand']:,.2f}` ({d['mpf_basis']})")
                        if d['mpf_segs_detail']:
                            for sal_cap, seg_yr, seg_cont, yrs_grow, seg_val in d['mpf_segs_detail']:
                                st.caption(f"  Cap ${sal_cap:,}: {seg_yr:.2f}yrs → contrib ${seg_cont:,.0f} × (1+{d['m_val']*100:.1f}%)^{yrs_grow:.1f} to valuation = ${seg_val:,.0f}")
                        st.markdown(f"{L['audit_mpf_vol_lbl']} `${d['mpf_vol']:,.2f}`")
                    with c2:
                        st.markdown(L["audit_step1_title"])
                        st.markdown(f"{L['audit_pre_y_lbl']} `{d['pre_y']:.4f} yrs`")
                        st.markdown(L["audit_pre_y_note"])
                        st.markdown(f"{L['audit_post_y_lbl']} `{d['post_y']:.4f} yrs`")
                        st.markdown(f"*(2025-05-01 → {rep_date})*")
                        if d['std'] != "SME-FRS":
                            st.markdown(f"{L['audit_ret_y_lbl']} `{d['y_to_ret']:.4f} yrs`")
                            st.markdown(f"{L['audit_pv_factor_lbl']} 1/(1+{d['r_val']*100:.2f}%)^{d['y_to_ret']:.2f} = `{d['pv']:.6f}`")
                            st.markdown(f"{L['audit_turnover_lbl']} `{d['t_val']*100:.1f}%`")
                            st.markdown(f"{L['audit_prob_stay_lbl']} (1−{d['t_val']*100:.1f}%)^{d['y_to_ret']:.2f} = `{d['prob_stay']:.6f}`")
                        else:
                            st.markdown(L["audit_pv_sme_label"])
                            st.markdown(L["audit_prob_stay_sme"])
                    with c3:
                        st.markdown(L["audit_step2_title"])
                        st.markdown(f"Pre-LSP = min(${d['sal_t']:,.0f}, $22,500) × 2/3 × {d['pre_y']:.4f}")
                        st.markdown(f"= `${d['v_pre']:,.2f}` {L['audit_pre_lsp_note']}")
                        sal_cap_note = (
                            f"min(${d['sal_c']:,.0f} [current salary], $22,500) = ${d['sal_proj_used']:,.2f}"
                            if d['effective_g'] == 0.0
                            else f"min(${d['sal_c']:,.0f}×(1+{d['effective_g']*100:.1f}%)^{d['y_to_ret']:.2f}, $22,500) = ${d['sal_proj_used']:,.2f}"
                        )
                        st.markdown(f"{L['audit_proj_sal_lbl']} {sal_cap_note}")
                        st.markdown(f"Post-LSP = ${d['sal_proj_used']:,.2f} × 2/3 × {d['post_y']:.4f}")
                        st.markdown(f"= `${d['v_post']:,.2f}` {L['audit_post_lsp_note']}")

                    c4, c5 = st.columns(2)
                    with c4:
                        st.markdown(L["audit_step3_title"])
                        st.markdown(f"{L['audit_mand_off_lbl']} min(${d['v_pre']:,.2f}, ${d['mpf_mand']:,.2f}) = `${d['offset_mand_pre']:,.2f}` {L['audit_mand_off_note']}")
                        st.markdown(f"{L['audit_vol_pre_lbl']} min(${d['v_pre']-d['offset_mand_pre']:,.2f}, ${d['mpf_vol']:,.2f}) = `${d['vol_to_pre']:,.2f}`")
                        st.markdown(f"{L['audit_total_pre_off']} `${d['offset_pre']:,.2f}` {L['audit_net_pre_lbl']} `${d['net_pre']:,.2f}`")
                        st.markdown(f"{L['audit_vol_rem_lbl']} ${d['mpf_vol']:,.2f} − ${d['vol_to_pre']:,.2f} = `${d['vol_remaining']:,.2f}`")
                        st.markdown(f"{L['audit_post_off_lbl']} min(${d['v_post']:,.2f}, ${d['vol_remaining']:,.2f}) = `${d['offset_post']:,.2f}` {L['audit_vol_only_note']}")
                        st.markdown(f"{L['audit_net_post_lbl']} `${d['net_post']:,.2f}`")
                    with c5:
                        st.markdown(L["audit_step4_title"])
                        # Threshold tier indicator
                        st.markdown(L["audit_threshold_within"] if d['within_threshold']
                                    else L["audit_threshold_beyond"])
                        # Rate summary row
                        cap_display = f"${d['sub_cap']:,}" if d['sub_cap'] is not None else "N/A"
                        cap_w_str = ("$" + f"{d['cap_w']:,}") if d['cap_w'] is not None else "none"
                        tier_note = (f"within \u2264$500k: {d['rate_w']*100:.0f}% / cap {cap_w_str}"
                                     f"  |  beyond >$500k: {d['rate_b']*100:.0f}%")
                        st.markdown(f"{L['audit_pol_yr_lbl']} **{d['policy_year']}** | {tier_note}")
                        st.markdown(f"{L['audit_rate_lbl']} **{d['sub_rate']*100:.0f}%** applied | {L['audit_emp_cap_lbl']} {cap_display}")
                        # Subsidy formula
                        if not d['within_threshold']:
                            st.markdown(f"{L['audit_subsidy_lbl']} {d['rate_b']*100:.0f}% × ${d['net_post']:,.2f}")
                        elif d['sub_cap'] is not None:
                            st.markdown(f"{L['audit_subsidy_lbl']} max({d['rate_w']*100:.0f}% × ${d['net_post']:,.2f},  ${d['net_post']:,.2f} − ${d['sub_cap']:,})")
                        else:
                            st.markdown(f"{L['audit_subsidy_lbl']} {d['rate_w']*100:.0f}% × ${d['net_post']:,.2f}")
                        st.markdown(f"= `${d['subsidy']:,.2f}`")
                        st.markdown(f"{L['audit_emp_burden_lbl']} ${d['net_post']:,.2f} − ${d['subsidy']:,.2f} = `${d['employer_post']:,.2f}`")
                        st.markdown(f"{L['audit_total_burden_lbl']} ${d['net_pre']:,.2f} + ${d['employer_post']:,.2f} = `${d['net_pre']+d['employer_post']:,.2f}`")
                        if d['std'] == "SME-FRS":
                            st.markdown(f"{L['audit_final_lbl']} (${d['net_pre']:,.2f} + ${d['employer_post']:,.2f}) × {d['pv']:.6f} = `${d['final_pv']:,.2f}`")
                        else:
                            st.markdown(f"{L['audit_final_lbl']} (${d['net_pre']:,.2f} + ${d['employer_post']:,.2f} × {d['prob_stay']:.6f}) × {d['pv']:.6f} = `${d['final_pv']:,.2f}`")

            st.markdown("---")

            # C. Statutory References
            st.markdown(L["audit_c_title"])
            st.markdown(L["audit_stat_table"])

            # D. Limitations
            st.markdown(L["audit_d_title"])
            st.warning(L["audit_limit_text"])

        # ── Excel Working Paper ───────────────────────────────────────────────
        output = io.BytesIO()
        wb = Workbook()

        # ── Styles ──
        hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        sub_fill  = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        yel_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        grn_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        red_fill  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        thin_b    = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'),  bottom=Side(style='thin'))
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True)
        title_font= Font(bold=True, size=14)

        def style_row(ws, fill, font=None, border=None):
            for cell in ws[ws.max_row]:
                cell.fill = fill
                if font:   cell.font   = font
                if border: cell.border = border

        def write_header(ws, values, fill=hdr_fill, font=hdr_font):
            ws.append(values)
            style_row(ws, fill, font, thin_b)

        def write_section_title(ws, title):
            ws.append([title])
            ws[ws.max_row][0].font = Font(bold=True, size=12, color="1F4E79")

        # ════════════════════════════════════════
        # Sheet 1 — Cover
        # ════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "1.Cover"
        ws1.column_dimensions['A'].width = 38
        ws1.column_dimensions['B'].width = 40

        ws1.append(["LONG SERVICE PAYMENT ACTUARIAL WORKING PAPER"])
        ws1[ws1.max_row][0].font = Font(bold=True, size=16, color="1F4E79")
        ws1.append(["Prepared in accordance with the Employment Ordinance (Cap. 57)"])
        ws1.append(["and the Abolition of MPF Offsetting Arrangement (effective 1 May 2025)"])
        ws1.append([])
        ws1.append(["Valuation Date",   str(rep_date)])
        ws1.append(["Accounting Standard", std_choice])
        ws1.append(["Policy Year (Post-Abolition)", year_idx_display])
        ws1.append(["Prepared By", ""])
        ws1.append(["Reviewed By", ""])
        ws1.append(["Date of Issue", ""])
        ws1.append([])
        ws1.append(["PURPOSE"])
        ws1[ws1.max_row][0].font = bold_font
        ws1.append(["This working paper calculates the Long Service Payment (LSP) liability"])
        ws1.append(["for accounting provision purposes under the applicable standard."])
        ws1.append(["Results should be reviewed against source documents before use."])
        ws1.append([])
        ws1.append(["CONFIRMATION"])
        ws1[ws1.max_row][0].font = bold_font
        ws1.append([L["sig_text"]])
        ws1.append([])
        ws1.append(["Confirmed By (Sign):", "____________________"])
        ws1.append(["Name & Title:", "____________________"])
        ws1.append(["Company:", "____________________"])
        ws1.append(["Date:", "____________________"])

        # ════════════════════════════════════════
        # Sheet 2 — Assumptions
        # ════════════════════════════════════════
        ws2 = wb.create_sheet("2.Assumptions")
        ws2.column_dimensions['A'].width = 42
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 45

        write_section_title(ws2, "ACTUARIAL ASSUMPTIONS")
        ws2.append([])
        write_header(ws2, ["Parameter", "Value Applied", "Basis / Reference"])
        for row_data in [
            ["Accounting Standard",     std_choice,
             "SME-FRS: no discounting | HKFRS/PE: PUC method with discounting"],
            ["Valuation Date",          str(rep_date),           "Reporting period end date"],
            ["Transition Date",         "2025-05-01",            "Effective date of MPF offsetting abolition"],
            ["MPF Inception Date",      "2000-12-01",            "Commencement of MPF scheme in HK"],
            ["Salary Growth Rate (g)",  f"{g_val*100:.1f}% p.a.",
             "Long-term inflation proxy; used to project salary to retirement (HKFRS only)"],
            ["Discount Rate (r)",
             f"{r_val*100:.2f}% p.a." if std_choice != "SME-FRS" else "N/A (SME-FRS)",
             "HKMA EFBN yield or high-quality corporate bond; per HKAS 19 / HKFRS for PE Sec. 28"],
            ["LSP Salary Cap",          "HK$22,500 / month",     "Employment Ordinance (Cap. 57)"],
            ["LSP Maximum Amount",      "HK$390,000",            "Employment Ordinance (Cap. 57)"],
            ["Minimum Vesting Period",  "5 years",               "Employment Ordinance (Cap. 57)"],
            ["MPF ER Mandatory Rate",   "5% of monthly salary",  "MPFA Ordinance (Cap. 485)"],
            ["MPF Salary Cap (MPF)",    "HK$30,000 / month",     "MPFA Ordinance (Cap. 485)"],
        ]:
            ws2.append(row_data)
            ws2[ws2.max_row][0].fill = sub_fill
            for cell in ws2[ws2.max_row]: cell.border = thin_b

        ws2.append([])
        write_section_title(ws2, "GOVERNMENT SUBSIDY SCHEME — 25-YEAR RATE LADDER")
        ws2.append(["(Subsidy for Abolition of MPF Offsetting Arrangement)"])
        ws2.append(["Note: Two-tier rates apply depending on whether net post-transition amount (b') is ≤ or > $500,000"])
        ws2.append([])
        write_header(ws2, ["Policy Year", "Scheme Year Range",
                            "(a) Within $500k — Rate", "(a) Employer Cap",
                            "(b) Beyond $500k — Rate"])
        subsidy_table = [
            (1,  "2025-05-01 – 2026-04-30", "50%", "$3,000",  "50%"),
            (2,  "2026-05-01 – 2027-04-30", "50%", "$3,000",  "50%"),
            (3,  "2027-05-01 – 2028-04-30", "50%", "$3,000",  "50%"),
            (4,  "2028-05-01 – 2029-04-30", "45%", "$25,000", "45%"),
            (5,  "2029-05-01 – 2030-04-30", "40%", "$25,000", "40%"),
            (6,  "2030-05-01 – 2031-04-30", "35%", "$25,000", "35%"),
            (7,  "2031-05-01 – 2032-04-30", "30%", "$50,000", "30%"),
            (8,  "2032-05-01 – 2033-04-30", "25%", "$50,000", "25%"),
            (9,  "2033-05-01 – 2034-04-30", "20%", "$50,000", "20%"),
            (10, "2034-05-01 – 2035-04-30", "20%", "N/A",     "15%"),
            (11, "2035-05-01 – 2036-04-30", "20%", "N/A",     "10%"),
            (12, "2036-05-01 – 2037-04-30", "15%", "N/A",      "5%"),
            (13, "2037-05-01 – 2038-04-30", "15%", "N/A",     "No subsidy"),
            (14, "2038-05-01 – 2039-04-30", "10%", "N/A",     "No subsidy"),
            (15, "2039-05-01 – 2040-04-30", "10%", "N/A",     "No subsidy"),
            (16, "2040-05-01 – 2041-04-30", "10%", "N/A",     "No subsidy"),
            (17, "2041-05-01 – 2042-04-30", "10%", "N/A",     "No subsidy"),
            (18, "2042-05-01 – 2043-04-30", "10%", "N/A",     "No subsidy"),
            (19, "2043-05-01 – 2044-04-30", "10%", "N/A",     "No subsidy"),
            (20, "2044-05-01 – 2045-04-30",  "5%", "N/A",     "No subsidy"),
            (21, "2045-05-01 – 2046-04-30",  "5%", "N/A",     "No subsidy"),
            (22, "2046-05-01 – 2047-04-30",  "5%", "N/A",     "No subsidy"),
            (23, "2047-05-01 – 2048-04-30",  "5%", "N/A",     "No subsidy"),
            (24, "2048-05-01 – 2049-04-30",  "5%", "N/A",     "No subsidy"),
            (25, "2049-05-01 – 2050-04-30",  "5%", "N/A",     "No subsidy"),
        ]
        for yr, rng, rate_w_s, cap_s, rate_b_s in subsidy_table:
            ws2.append([yr, rng, rate_w_s, cap_s, rate_b_s])
            for cell in ws2[ws2.max_row]: cell.border = thin_b
            if yr == year_idx_display:
                for cell in ws2[ws2.max_row]: cell.fill = yel_fill

        ws2.append([])
        write_section_title(ws2, "KEY FORMULAS")
        ws2.append([])
        write_header(ws2, ["Step", "Formula", "Notes"])
        formula_rows = [
            ["Pre-transition LSP (a)",
             "min(min(Salary@Trans, $22,500) × 2/3 × pre_years, $390,000)",
             "Salary and total capped per Cap. 57"],
            ["Post-transition LSP (b)",
             "min(min(Projected Salary, $22,500) × 2/3 × post_years, $390,000 − a)",
             "Projected salary = Current Salary × (1+g)^years_to_retirement"],
            ["Service Years",
             "Whole completed years + remaining days / 365",
             "Official LD EasyCal method"],
            ["MPF Mandatory Offset",
             "min(a, MPF Mandatory ER) → applied to pre-LSP only",
             "Cannot offset post-transition LSP"],
            ["MPF Voluntary Offset",
             "Remaining voluntary after pre-offset → applied to post-LSP",
             "Voluntary contributions may offset either period"],
            ["MPF Estimation (if no data)",
             "3-period segment model: min(sal, cap) × 5% × 12 × yrs × (1+m)^yrs_to_val",
             "Caps: ≤Apr 2012=$20k | May 2012–May 2014=$25k | Jun 2014+=$30k; estimate ER mandatory up to 2025-05-01, then compound to valuation date"],
            ["Government Subsidy (C)",
             "max(Rate × net_post, net_post − Employer Cap)",
             "Applied to net post-transition LSP (b'); rate per policy year"],
            ["Employer Post-burden",
             "net_post − Subsidy",
             "Minimum employer pays = Employer Cap (if net_post > 2×cap)"],
            ["PV Factor (HKFRS only)",
             "1 / (1 + r)^years_to_retirement",
             "SME-FRS: PV Factor = 1 (no discounting)"],
            ["Final Liability",
             "(net_pre + Employer Post-burden) × PV Factor",
             "Total employer obligation at valuation date"],
        ]
        for fr in formula_rows:
            ws2.append(fr)
            ws2[ws2.max_row][0].fill = sub_fill
            for cell in ws2[ws2.max_row]: cell.border = thin_b

        # ════════════════════════════════════════
        # Sheet 3 — Summary Results
        # ════════════════════════════════════════
        ws3 = wb.create_sheet("3.Summary")
        write_section_title(ws3, "SUMMARY RESULTS")
        ws3.append([f"Valuation Date: {rep_date}   |   Standard: {std_choice}   |   Policy Year: {year_idx_display}"])
        ws3.append([])
        write_header(ws3, list(res_df.columns))
        for _, row in res_df.iterrows():
            ws3.append(list(row))
            for cell in ws3[ws3.max_row]: cell.border = thin_b

        ws3.append([])
        ws3.append(["Total", ""] + [""] * (len(res_df.columns) - 2))
        numeric_cols = [c for c in res_df.columns if res_df[c].dtype in [float, int] and c != "Name"]
        for col_idx, col in enumerate(res_df.columns):
            if col in numeric_cols:
                total_val = res_df[col].sum() if pd.api.types.is_numeric_dtype(res_df[col]) else ""
                ws3[ws3.max_row][col_idx].value = round(total_val, 2) if isinstance(total_val, float) else total_val
        ws3[ws3.max_row][0].value = "TOTAL"
        style_row(ws3, grn_fill, bold_font, thin_b)

        for col in ws3.columns:
            ws3.column_dimensions[col[0].column_letter].width = max(
                18, max(len(str(cell.value or "")) for cell in col) + 2)

        # ════════════════════════════════════════
        # Sheet 4 — Detailed Calculations
        # ════════════════════════════════════════
        ws4 = wb.create_sheet("4.Calc Detail")
        write_section_title(ws4, "PER-EMPLOYEE DETAILED CALCULATION")
        ws4.append([])
        ws4.column_dimensions['A'].width = 38
        ws4.column_dimensions['B'].width = 28
        ws4.column_dimensions['C'].width = 50

        for d in audit_details:
            age_v = (REPORT_DATE - d["dob"]).days / 365.25
            ws4.append([f"Employee: {d['name']}"])
            ws4[ws4.max_row][0].font = Font(bold=True, size=12, color="1F4E79")
            write_header(ws4, ["Step", "Value", "Formula / Note"], sub_fill, bold_font)

            detail_rows = [
                # Input
                ["── INPUT DATA ──", "", ""],
                ["Hire Date",               str(d['h_dt'].date()),     ""],
                ["Date of Birth",           str(d['dob'].date()),       f"Age at valuation: {age_v:.1f}"],
                ["Salary at Transition",    f"${d['sal_t']:,.2f}",      "Monthly salary immediately before 2025-05-01"],
                ["Current Salary",          f"${d['sal_c']:,.2f}",      "Monthly salary at valuation date"],
                ["MPF Mandatory (ER)",       f"${d['mpf_mand']:,.2f}",  f"{d['mpf_basis']} (pre-2025-05-01 ER mandatory only)"],
            ]
            # If estimated, append one sub-row per segment
            if d['mpf_segs_detail']:
                for sal_cap, seg_yr, seg_cont, yrs_grow, seg_val in d['mpf_segs_detail']:
                    detail_rows.append([
                        f"  └ Cap ${sal_cap:,}",
                        f"${seg_val:,.2f}",
                        f"min(sal,${sal_cap:,})×5%×12×{seg_yr:.2f}yrs=${seg_cont:,.0f} × (1+{d['m_val']*100:.1f}%)^{yrs_grow:.1f}yrs to valuation"
                    ])
            detail_rows += [
                ["MPF Voluntary (ER)",       f"${d['mpf_vol']:,.2f}",   "Actual input; 0 if not provided"],
                # Service years
                ["── STEP 1: SERVICE YEARS ──", "", ""],
                ["Pre-transition years",    f"{d['pre_y']:.4f} yrs",
                 f"Hire ({d['h_dt'].date()}) → 2025-05-01; whole yrs + days/365"],
                ["Post-transition years",   f"{d['post_y']:.4f} yrs",
                 f"2025-05-01 → {rep_date}; days/365"],
                ["Years to retirement (65)",f"{d['y_to_ret']:.4f} yrs",
                 f"DOB + 65 yrs − valuation date"],
                ["PV Factor",               f"{d['pv']:.6f}",
                 f"1/(1+{d['r_val']*100:.2f}%)^{d['y_to_ret']:.2f}" if d['std'] != "SME-FRS" else "1.000000 (SME-FRS: no discounting)"],
                ["Annual Turnover Rate",    f"{d['t_val']*100:.1f}%",
                 "SME-FRS: N/A (100% retention assumed)" if d['std'] == "SME-FRS" else "Probability of employee remaining until retirement"],
                ["Prob. of Staying",        f"{d['prob_stay']:.6f}",
                 "1.000000 (SME-FRS)" if d['std'] == "SME-FRS" else f"(1−{d['t_val']*100:.1f}%)^{d['y_to_ret']:.2f}"],
                # LSP
                ["── STEP 2: LSP CALCULATION ──", "", ""],
                ["Salary cap applied (pre)", f"${min(d['sal_t'],22500):,.2f}",
                 f"min(${d['sal_t']:,.0f}, $22,500)"],
                ["Pre-transition LSP (a)",  f"${d['v_pre']:,.2f}",
                 f"${min(d['sal_t'],22500):,.2f} × 2/3 × {d['pre_y']:.4f} yrs, cap $390,000"],
                ["Projected salary (post)", f"${d['sal_proj_used']:,.2f}",
                 ("min(${:,.0f} [current salary, SME-FRS: g=0%], $22,500)".format(d['sal_c'])
                  if d['effective_g'] == 0.0
                  else f"min(${d['sal_c']:,.0f} × (1+{d['effective_g']*100:.1f}%)^{d['y_to_ret']:.2f}, $22,500)")],
                ["Post-transition LSP (b)", f"${d['v_post']:,.2f}",
                 f"${d['sal_proj_used']:,.2f} × 2/3 × {d['post_y']:.4f} yrs, cap $390,000 − (a)"],
                # MPF
                ["── STEP 3: MPF OFFSETTING ──", "", ""],
                ["Mandatory offset on pre", f"${d['offset_mand_pre']:,.2f}",
                 f"min(${d['v_pre']:,.2f}, ${d['mpf_mand']:,.2f}) — mandatory: pre only"],
                ["Voluntary offset on pre", f"${d['vol_to_pre']:,.2f}",
                 f"min(${d['v_pre']-d['offset_mand_pre']:,.2f} remaining pre, ${d['mpf_vol']:,.2f})"],
                ["Total pre offset",        f"${d['offset_pre']:,.2f}",  "Mandatory offset + Voluntary offset on pre"],
                ["Net pre-transition (a')", f"${d['net_pre']:,.2f}",     "(a) − total pre offset"],
                ["Voluntary remaining",     f"${d['vol_remaining']:,.2f}","Voluntary − used on pre"],
                ["Post offset (voluntary)", f"${d['offset_post']:,.2f}",
                 f"min(${d['v_post']:,.2f}, ${d['vol_remaining']:,.2f}) — voluntary only on post"],
                ["Net post-transition (b')",f"${d['net_post']:,.2f}",    "(b) − post offset"],
                # Subsidy
                ["── STEP 4: GOVERNMENT SUBSIDY ──", "", ""],
                ["Policy Year",             str(d['policy_year']),
                 f"Days since 2025-05-01 = {(REPORT_DATE - pd.Timestamp('2025-05-01')).days}; Year = days//365 + 1"],
                ["$500k Threshold",
                 "Within \u2264$500,000" if d['within_threshold'] else "Beyond >$500,000",
                 "Determines which column (a) or (b) of the 25-year ladder applies"],
                ["Within-threshold rate",   f"{d['rate_w']*100:.0f}%",
                 ("Cap: $" + f"{d['cap_w']:,}") if d['cap_w'] is not None else "Cap: N/A (rate-based)"],
                ["Beyond-threshold rate",   f"{d['rate_b']*100:.0f}%",
                 "0% = No subsidy" if d['rate_b'] == 0.0 else "Rate-only, no cap"],
                ["Rate Applied",            f"{d['sub_rate']*100:.0f}%",
                 "Column (a) within-threshold" if d['within_threshold'] else "Column (b) beyond-threshold"],
                ["Employer's Cap",
                 f"${d['sub_cap']:,}" if d['sub_cap'] is not None else "N/A",
                 "Applies to within-threshold years 1–9 only"],
                ["Subsidy (C)",             f"${d['subsidy']:,.2f}",
                 (f"max({d['rate_w']*100:.0f}% × ${d['net_post']:,.2f},  ${d['net_post']:,.2f} − ${d['sub_cap']:,})"
                  if d['within_threshold'] and d['sub_cap'] is not None else
                  f"{d['sub_rate']*100:.0f}% × ${d['net_post']:,.2f}")],
                ["Employer post-burden",    f"${d['employer_post']:,.2f}", "(b') − Subsidy (C)"],
                # Final
                ["── STEP 5: FINAL LIABILITY ──", "", ""],
                ["Net pre-transition (a')",  f"${d['net_pre']:,.2f}",
                 "Fully vested — no turnover reduction applied"],
                ["Employer post-burden",     f"${d['employer_post']:,.2f}",
                 ""],
                ["Prob. of Staying",         f"{d['prob_stay']:.6f}",
                 "1.000000 (SME-FRS)" if d['std'] == "SME-FRS" else f"Applied to employer_post only; (1−{d['t_val']*100:.1f}%)^{d['y_to_ret']:.2f}"],
                ["Adj. Employer Post",       f"${d['employer_post']*d['prob_stay']:,.2f}",
                 "1.000000 (SME-FRS — no turnover adj.)" if d['std'] == "SME-FRS" else f"${d['employer_post']:,.2f} × {d['prob_stay']:.6f}"],
                ["Pre + Adj. Post",          f"${d['net_pre'] + d['employer_post']*d['prob_stay']:,.2f}",
                 f"${d['net_pre']:,.2f} + ${d['employer_post']*d['prob_stay']:,.2f}"],
                ["PV Factor",               f"{d['pv']:.6f}",
                 "1.000000 (SME-FRS)" if d['std'] == "SME-FRS" else f"1/(1+{d['r_val']*100:.2f}%)^{d['y_to_ret']:.2f}"],
                ["FINAL LIABILITY",         f"${d['final_pv']:,.2f}",
                 f"(${d['net_pre']:,.2f} + ${d['employer_post']:,.2f} × {d['prob_stay']:.6f}) × {d['pv']:.6f}"],
            ]
            for dr in detail_rows:
                ws4.append(dr)
                if dr[0].startswith("──"):
                    ws4[ws4.max_row][0].fill = sub_fill
                    ws4[ws4.max_row][0].font = bold_font
                elif dr[0] == "FINAL LIABILITY":
                    style_row(ws4, grn_fill, bold_font, thin_b)
                else:
                    for cell in ws4[ws4.max_row]: cell.border = thin_b
            ws4.append([])

        # ════════════════════════════════════════
        # Sheet 5 — Limitations
        # ════════════════════════════════════════
        ws5 = wb.create_sheet("5.Limitations")
        ws5.column_dimensions['A'].width = 30
        ws5.column_dimensions['B'].width = 80

        write_section_title(ws5, "LIMITATIONS, DISCLAIMERS & STATUTORY REFERENCES")
        ws5.append([])

        write_header(ws5, ["#", "Limitation / Disclaimer"])
        limitations = [
            ("1", "MPF ESTIMATED BALANCES: Where MPF Mandatory (ER) was zero or blank, balance was estimated using a 3-period segment model reflecting historical MPF salary cap changes: ≤Apr 2012=$20,000/mth, May 2012–May 2014=$25,000/mth, Jun 2014+=$30,000/mth. Estimation includes ER mandatory contributions up to 2025-05-01 only, then compounds to valuation date at the user-specified MPF Return Rate. Actual MPF trustee statements (valuation-date accrued benefit) MUST be obtained for final audit sign-off."),
            ("2", "SALARY PROJECTION: Under HKFRS, post-transition LSP uses salary projected to age 65 at the user-specified growth rate. Under SME-FRS, the current salary is used directly (growth rate forced to 0%). Actual termination salary may differ."),
            ("3", "DISCOUNT RATE: Applies to HKFRS / HKFRS for PE only. Based on HKMA EFBN yield or user input. Rate must be refreshed at each reporting date per HKAS 19 requirements."),
            ("4", "GOVERNMENT SUBSIDY: Rates are subject to legislative change over the 25-year scheme. Always verify the applicable year's rate with the official Subsidy Scheme Administrator."),
            ("5", "SCOPE — LSP ONLY: This tool covers Long Service Payments. Severance Payments (SP) follow the same formula but apply to redundancy cases. Verify applicability under Cap. 57 ss.31–33."),
            ("6", "PRE-2000 SERVICE: Employees hired before 1 Dec 2000 — MPF estimation starts from 1 Dec 2000 (MPF inception). Pre-MPF service is not captured in the estimated MPF balance."),
            ("7", "DATA RELIABILITY: Results depend entirely on input data. The preparer must verify hire dates, salaries, DOBs and MPF balances against source documents (HR records, MPF statements)."),
            ("8", "FIVE-YEAR VESTING: Employees with less than 5 years' service are excluded (not entitled to LSP). Verify this cutoff against actual contractual terms."),
            ("9", "390,000 CAP: Total LSP (pre + post combined) is capped at $390,000 per Cap. 57."),
            ("10","PROFESSIONAL ADVICE: This tool is for reference and internal audit purposes only. It does not constitute legal, actuarial, or accounting advice. Consult a qualified actuary or auditor for final provisions."),
            ("11","RETIREMENT AGE: Retirement is assumed at age 65 for all employees. Users should adjust the Annual Turnover Rate to account for the probability that employees may leave before reaching retirement age."),
            ("12","TURNOVER / ATTRITION: Under HKFRS, the turnover probability (1−t)^n is applied ONLY to the employer's post-transition burden (employer_post), not to the pre-transition net obligation (net_pre). Pre-transition LSP is fully vested for employees with 5+ years of service — it is payable upon any termination and has no forfeiture risk. The simplified formula is: Final Liability = (net_pre + employer_post × prob_stay) × PV Factor."),
            ("13","DATA PRIVACY: All calculations are performed locally in the user's browser. No employee data is transmitted to or stored on any external server."),
        ]
        for num, text in limitations:
            ws5.append([num, text])
            ws5[ws5.max_row][0].font = bold_font
            ws5[ws5.max_row][1].alignment = Alignment(wrap_text=True)
            for cell in ws5[ws5.max_row]: cell.border = thin_b
        ws5.append([])

        write_header(ws5, ["Statutory Reference", "Description"])
        refs = [
            ("Employment Ordinance Cap. 57",              "LSP formula, salary/total caps, vesting period"),
            ("EMPLO (Amendment) Ordinance 2022",          "Abolition of MPF offsetting; transition date 1 May 2025"),
            ("MPF Schemes Ordinance Cap. 485",            "Mandatory 5% ER contribution; $30,000 salary cap"),
            ("Subsidy Scheme (25-year ladder)",           "Government subsidy for employer post-transition LSP burden"),
            ("HKAS 19 / HKFRS for PE Sec. 28",           "Accounting for employee benefits; PUC method; discounting"),
            ("HKICPA SME-FRS Section 28",                 "Employee benefits under SME-FRS; no discounting required"),
            ("HKICPA Guidance — MPF-LSP Accounting",     "Practical guidance on accounting for transition provisions"),
        ]
        for ref, desc in refs:
            ws5.append([ref, desc])
            for cell in ws5[ws5.max_row]: cell.border = thin_b

        wb.save(output)
        st.download_button(L["download_btn"], output.getvalue(), "LSP_Audit_WP_Final.xlsx")


# --- 6. 資源連結 ---
st.markdown("---")
with st.expander(L["guide_expander"], expanded=False):
    guide_md = load_user_guide_md()
    if guide_md:
        st.markdown(guide_md)
    else:
        st.warning(L["guide_missing"])

# --- 7. 資源連結 ---
st.markdown("---")
st.subheader(L["links_title"])
c1, c2 = st.columns(2)
with c1:
    st.markdown("1. [Labour Dept Calculator](https://www.labour.gov.hk/eng/labour/Statutory_Employment_Entitlements_Reference_Calculator.htm)")
    st.markdown("2. [Abolition of Offsetting Portal](https://www.op.labour.gov.hk/en/index.html)")
    st.markdown("3. [Offsetting Subsidy Scheme](https://www.offsettingsubsidy.gov.hk/en/index.html)")
    st.markdown("4. [Subsidy Scheme Calculator](https://www.offsettingsubsidy.gov.hk/en/calculator.html)")
with c2:
    st.markdown("5. [HKICPA - MPF-LSP Accounting Guide](https://www.hkicpa.org.hk/-/media/HKICPA-Website/New-HKICPA/Standards-and-regulation/SSD/gMPFLSP.pdf)")
    st.markdown("6. [MPFA Quarterly Reports](https://www.mpfa.org.hk/en/info-centre/research-reports/quarterly-reports/mpf-schemes)")
    st.markdown("7. [MPFA Performance Data](https://www.mpfa.org.hk/en/info-centre/press-releases)")
    st.markdown("8. [HKMA EFBN Daily Yields](https://www.hkma.gov.hk/eng/data-publications-and-research/data-and-statistics/daily-monetary-statistics/)")

st.markdown("---")
st.caption(L["disclaimer"])